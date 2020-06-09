from typing import Dict, List
import peewee
from openpyxl import Workbook, load_workbook
from lib import scryfall, logging
from models import Card


log = logging.get(name='card_service')


def create_card(card_json: Dict) -> Card:
    params = dict()
    for field in Card._meta.fields.keys():
        params[field] = card_json.get(field, None)

    try:
        card = Card.create(**params)
    except peewee.DatabaseError as e:
        log.warning("Unable to create card with the following params:")
        log.warning(params)
        log.warning(e)
        return None

    log.info("Created card [%s]", card.name)

    return card


def get_card(card_name: str) -> Card:
    try:
        card = Card.get(Card.name == card_name)
    except Card.DoesNotExist:
        log.debug('Card not found, searching Scryfall')
        card_json = scryfall.get_card_exact(card_name)
        card = create_card(card_json)

    return card


# TODO: Auto-size columns
def write_to_workbook(deck_name: str, cards: List[Card], path: str) -> None:
    try:
        workbook = load_workbook(filename=path)
    except FileNotFoundError:
        log.info("File not found, creating new one")
        workbook = Workbook()

        # Remove default worksheet
        worksheet = workbook.active
        workbook.remove(worksheet)

    worksheet = workbook.create_sheet(title=deck_name)

    for idx, card in enumerate(cards):
        worksheet.cell(row=idx+1, column=1, value=card.name)
        worksheet.cell(row=idx+1, column=2, value=card.token_type)
        worksheet.cell(row=idx+1, column=3, value=card.oracle_text)

    workbook.save(filename=path)
